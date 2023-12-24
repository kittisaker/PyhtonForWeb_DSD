from openpyxl import load_workbook
import matplotlib.pyplot as plt
import datetime 

workbook = load_workbook(filename="data.xlsx")
sheet = workbook.active

x =[]
for i, col in enumerate(sheet):
    if i == 0:
        continue
    data_get = col[0].value
    x.append(int(data_get))
#print(x)

y =[]
for i, col in enumerate(sheet):
    if i == 0:
        continue
    data_get = col[1].value
    y.append(int(data_get))
#print(y)
    
max_col = sheet.max_column
list_head=[ ]
for i in range(1, max_col+1):
    cell_obj = sheet.cell(row = 1, column= i)
    list_head.append(cell_obj.value)
    

plt.plot(x,y)
plt.xlabel(list_head[0]) 
plt.ylabel(list_head[1]) 
now = datetime.datetime.now()
plt.title(now.strftime("%Y-%m-%d %H:%M:%S"))
plt.show()