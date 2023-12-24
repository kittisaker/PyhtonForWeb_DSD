from openpyxl import load_workbook

workbook = load_workbook(filename="data.xlsx")
# print(workbook.sheetnames)

sheet = workbook.active
# print(sheet.title)
# print(sheet.cell(row=1, column=1))  # <Cell 'Sheet1'.A1>
# print(sheet.cell(row=1, column=1).value)    #price

# --------------------------------------------------
# for i in range(1, 12):
#     print(sheet.cell(row=i, column=1).value)

# --------------------------------------------------
# add to list
# x = []

# for row in sheet:
#     data_get = row[0].value
#     x.append(data_get)

# print(x)

# --------------------------------------------------
# 
x =  []
for i, col in enumerate(sheet):
    if i == 0:
        continue
    data_get = col[0].value
    x.append(int(data_get))
print(x)

y =  []
for i, col in enumerate(sheet):
    if i == 0:
        continue
    data_get = col[1].value
    y.append(int(data_get))
print(y)

# --------------------------------------------------