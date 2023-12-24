#python-excel
import openpyxl
wb_obj = openpyxl.load_workbook("data.xlsx")
sheet_obj = wb_obj.active

#scan "text" in colum
max_col = sheet_obj.max_column
list_head=[ ]
for i in range(1, max_col+1):
    cell_obj = sheet_obj.cell(row = 1, column= i)
    list_head.append(cell_obj.value)
print(list_head)
    
#scan "text" in row
max_row = sheet_obj.max_row
list_rowA=[ ]
for i in range(1, max_row+1):
    cell_obj =sheet_obj.cell(row = i , column = 1)
    list_rowA.append(cell_obj.value)
print(list_rowA)

#scan "text" in row
max_row = sheet_obj.max_row
list_rowF=[ ]
for i in range(1, max_row+1):
    cell_obj =sheet_obj.cell(row = i , column = 2)
    list_rowF.append(cell_obj.value)
print(list_rowF)